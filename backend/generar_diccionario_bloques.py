import os
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Configuración de conexión a PostgreSQL
DB_CONFIG = {
    'dbname': 'bdexpo',
    'user': 'postgres',
    'password': 'root',
    'host': 'localhost',
    'port': '5432'
}

# Definición de Módulos y sus Tablas correspondientes en orden
MODULOS = {
    "Seguridad e Identidad": [
        "usuario", "cargo", "personal", "roles", "permiso", "roles_permiso", "roles_permiso_usuario", 
        "usuario_groups", "usuario_user_permissions"
    ],
    "Académico": [
        "entidad_academica", "carrera", "ea_carrera", "area", "modalidad", "modalidad_area", "oferta", "oferta_ea_carrera", "categoria"
    ],
    "Eventos y Logística": [
        "evento", "nivel_evento", "categoria_evento", "tipo_evento", "membrete", "membrete_firmante", "cronograma", "actividad"
    ],
    "Proyectos y Registro": [
        "proyecto", "participante", "tutor", "proyecto_participantes", "proyecto_tutores"
    ],
    "Evaluaciones": [
        "planilla_evaluativa", "seccion", "criterio", "acta_evaluacion", "detalle_evaluacion", "puntuacion_criterio", 
        "tribunal", "tribunal_ext", "tribunal_areas", "proyecto_tribunales"
    ],
    "Premiación y Certificados": [
        "premio", "candidato_premio", "ganador_premio", "asignacion_premio", "descriptor", "tipo_descriptor", 
        "premio_descriptor", "plantilla", "certificado"
    ]
}

# Descripciones de cada tabla
TABLE_DESCRIPTIONS = {
    "usuario": "Almacena las credenciales de acceso, información de autenticación y claves TOTP para el doble factor (2FA) de los usuarios del sistema.",
    "cargo": "Define los cargos o puestos de las autoridades (ej. Rector, Decano) autorizadas para firmar certificados oficializados.",
    "personal": "Registra al personal administrativo y autoridades universitarias que intervienen como firmantes autorizados del evento.",
    "roles": "Define los roles del sistema (Administrador, Tribunal, Participante, etc.) asignados a los usuarios.",
    "permiso": "Almacena los permisos específicos de acceso a vistas y funciones en la plataforma.",
    "roles_permiso": "Tabla intermedia que asocia los permisos habilitados para cada rol del sistema.",
    "roles_permiso_usuario": "Asociación directa de permisos adicionales o específicos a un usuario en particular.",
    "usuario_groups": "Tabla intermedia generada por Django para asociar usuarios a grupos de seguridad.",
    "usuario_user_permissions": "Tabla intermedia generada por Django para asociar permisos individuales de Django a usuarios.",
    
    "entidad_academica": "Registra las facultades de la UAGRM participantes en el evento (ej. Facultad de Ciencias de la Computación).",
    "carrera": "Registra las carreras académicas ofertadas en la universidad (nombre, plan de estudios y código descriptivo).",
    "ea_carrera": "Tabla intermedia que relaciona cada facultad (entidad académica) con sus respectivas carreras académicas.",
    "area": "Registra las áreas científicas y tecnológicas en las que se clasifican los proyectos (ej. Desarrollo de Software).",
    "modalidad": "Define las modalidades de participación (ej. Feria Científica, Proyecto de Grado).",
    "modalidad_area": "Tabla intermedia que vincula qué áreas académicas están disponibles para cada modalidad.",
    "oferta": "Define las ofertas de categorías académicas activas para el evento.",
    "oferta_ea_carrera": "Vincula la oferta académica con una carrera y facultad específica para la postulación.",
    "categoria": "Registra las categorías de participación de los estudiantes.",
    
    "evento": "Registra los eventos o ferias científicas de Expociencia creados en el sistema (nombre, versión, año).",
    "nivel_evento": "Define el alcance del evento (Facultativo, Interfacultativo o Universitario).",
    "categoria_evento": "Registra las categorías asignadas a un evento específico.",
    "tipo_evento": "Registra las clasificaciones o tipos de eventos.",
    "membrete": "Almacena la configuración visual de logos, sellos y encabezados para los certificados oficiales.",
    "membrete_firmante": "Tabla intermedia que asocia a las autoridades del personal administrativo (firmantes) a un membrete y define el orden de aparición de sus firmas.",
    "cronograma": "Define el calendario y fechas límite para cada actividad del evento (ej. período de inscripción).",
    "actividad": "Registra las actividades individuales asociadas al cronograma de un evento.",
    
    "proyecto": "Almacena la información principal de los proyectos postulados (título, resumen, archivo digital y estado).",
    "participante": "Registra los perfiles de los estudiantes integrantes de los proyectos, enlazados a su código de registro universitario.",
    "tutor": "Registra a los docentes que actúan como tutores o guías metodológicos de los proyectos.",
    "proyecto_participantes": "Tabla de relación muchos a muchos entre los proyectos y los estudiantes integrantes.",
    "proyecto_tutores": "Tabla de relación muchos a muchos entre los proyectos y sus respectivos docentes tutores.",
    
    "planilla_evaluativa": "Define las rúbricas o planillas de evaluación creadas por los administradores para calificar los proyectos.",
    "seccion": "Registra las secciones ponderadas dentro de una planilla evaluativa (ej. Exposición oral, Documento escrito).",
    "criterio": "Registra los criterios y puntajes individuales de calificación pertenecientes a una sección evaluativa.",
    "acta_evaluacion": "Consolida las notas finales de cada proyecto en base al promedio de las evaluaciones recibidas.",
    "detalle_evaluacion": "Registra las calificaciones completas enviadas por cada tribunal asignado a un proyecto.",
    "puntuacion_criterio": "Almacena los puntos específicos otorgados por un jurado en un criterio evaluativo particular.",
    "tribunal": "Registra a los profesionales y docentes asignados como jurados evaluadores.",
    "tribunal_ext": "Registra datos de jurados evaluadores externos a la universidad (ej. institución u organización a la que representan).",
    "tribunal_areas": "Vincula a los jurados con sus áreas de especialidad para evitar asignaciones incorrectas.",
    "proyecto_tribunales": "Tabla intermedia que asocia a los jurados (tribunales) con los proyectos que deben calificar.",
    
    "premio": "Registra los premios configurados para el evento (ej. Primer Lugar, Mención de Honor).",
    "candidato_premio": "Registra los proyectos postulados que quedan en estado de empate técnico durante el cierre de actas.",
    "ganador_premio": "Registra de forma oficial los proyectos declarados como ganadores de los premios.",
    "asignacion_premio": "Gestiona la asignación y desglose de los montos financieros o trofeos a los ganadores.",
    "descriptor": "Define las descripciones breves asociadas a las categorías de premios.",
    "tipo_descriptor": "Define las tipologías para las descripciones de premios.",
    "premio_descriptor": "Vincula los premios con descriptores específicos.",
    "plantilla": "Define las plantillas HTML configuradas para renderizar los certificados.",
    "certificado": "Almacena el registro único de certificados emitidos para estudiantes, tutores y jurados."
}

def generar_descripcion_campo(tabla, campo):
    # Diccionario de campos comunes
    campos_comunes = {
        "id": "Identificador único y autoincremental de la tabla (Llave Primaria).",
        "created_at": "Fecha y hora de creación automática del registro.",
        "updated_at": "Fecha y hora de la última modificación del registro.",
        "nombre": "Nombre descriptivo o denominación oficial.",
        "estado": "Estado actual del registro dentro de su flujo de trabajo.",
        "descripcion": "Descripción detallada del registro o notas adicionales.",
        "email": "Correo electrónico de contacto o institucional.",
        "telefono": "Número telefónico o de celular para comunicación.",
        "celular": "Número de teléfono celular.",
        "fecha": "Fecha asociada al registro.",
        "ci": "Cédula de Identidad de la persona.",
        "expedido": "Lugar de expedición del documento de identidad.",
        "codigo": "Código alfanumérico identificador.",
        "observacion": "Observaciones o comentarios adicionales sobre el registro.",
        "failed_login_attempts": "Cantidad de intentos de inicio de sesión fallidos consecutivos antes del bloqueo.",
        "locked_until": "Fecha y hora hasta la cual el usuario se encuentra bloqueado debido a intentos fallidos.",
        "firma_img": "Imagen digitalizada de la firma manuscrita de la autoridad.",
        "estado_pago": "Estado de la transferencia o pago del premio (ej. pendiente, pagado, rechazado).",
        "comprobante_pago_imagen": "Imagen de comprobante o captura de pantalla del pago por QR realizado.",
        "consolidada": "Indica si el acta de evaluación ya ha sido cerrada y consolidada de forma definitiva.",
        "desempate_prioridad": "Número de prioridad o puesto asignado en el proceso de resolución de desempates.",
        "orden": "Número de orden o secuencia de aparición de los registros."
    }
    
    if campo in campos_comunes:
        return campos_comunes[campo]
        
    # Claves foráneas genéricas
    if campo.startswith("id_") or campo.endswith("_id"):
        tabla_rel = campo.replace("id_", "").replace("_id", "")
        return f"Relación o clave foránea con la tabla de {tabla_rel}."
        
    if "fecha" in campo or "date" in campo:
        return "Fecha asociada al registro de la actividad."
        
    if "monto" in campo or "premio" in campo or "valor" in campo:
        return "Valor numérico o financiero asociado."
        
    # Descripciones específicas por tabla
    especificos = {
        "usuario": {
            "username": "Nombre de usuario para el inicio de sesión y autenticación.",
            "password": "Contraseña hasheada con algoritmos seguros del framework Django.",
            "is_active": "Indica si la cuenta del usuario está activa en el sistema.",
            "is_staff": "Indica si el usuario tiene privilegios para acceder al panel administrativo.",
            "is_superuser": "Indica si el usuario cuenta con permisos absolutos en toda la plataforma.",
            "dos_factor_key": "Clave secreta TOTP codificada para la verificación de doble factor (2FA).",
            "celular_verificado": "Indica si el teléfono celular del usuario ha sido verificado.",
            "last_login": "Fecha y hora del último inicio de sesión del usuario."
        },
        "proyecto": {
            "titulo": "Título formal y descriptivo del proyecto de investigación o feria.",
            "resumen": "Resumen técnico o abstract detallado de los objetivos del proyecto.",
            "archivo": "Ruta física o enlace al documento en formato PDF/Word del proyecto.",
            "nota_final": "Nota final calculada en base al promedio de las evaluaciones.",
            "observaciones": "Observaciones técnicas o comentarios del comité de revisión."
        },
        "participante": {
            "registro": "Código de registro universitario del estudiante en la UAGRM.",
            "semestre": "Semestre actual en el que se encuentra inscrito el estudiante.",
            "carrera": "Carrera de filiación del estudiante participante."
        },
        "tutor": {
            "registro_docente": "Código de registro docente oficial del tutor académico."
        },
        "cronograma": {
            "fecha_inicio": "Fecha y hora de inicio de la actividad o período.",
            "fecha_fin": "Fecha y hora de finalización de la actividad o período."
        },
        "criterio": {
            "puntaje_maximo": "Puntaje máximo de calificación permitido para este criterio."
        },
        "seccion": {
            "ponderacion": "Porcentaje de ponderación de la sección sobre la nota final (0-100%)."
        }
    }
    
    if tabla in especificos and campo in especificos[tabla]:
        return especificos[tabla][campo]
        
    return f"Campo técnico {campo} registrado en la tabla {tabla}."

def obtener_metadatos_base_datos():
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        print(f"Error de conexión: {e}")
        return None

    # Obtener todas las tablas en la base de datos
    cur.execute("""
        SELECT table_name 
        FROM information_schema.tables 
        WHERE table_schema = 'public' 
          AND table_type = 'BASE TABLE'
    """)
    tablas_existentes = set(r[0] for r in cur.fetchall())

    # Query para extraer las columnas detalladas
    query_columnas = """
        SELECT 
            c.column_name,
            c.data_type,
            c.character_maximum_length,
            c.numeric_precision,
            c.numeric_scale,
            c.is_nullable,
            c.column_default,
            -- PK
            (SELECT count(*) > 0 
             FROM information_schema.table_constraints tc
             JOIN information_schema.key_column_usage kcu ON tc.constraint_name = kcu.constraint_name
             WHERE tc.table_schema = c.table_schema 
               AND tc.table_name = c.table_name 
               AND kcu.column_name = c.column_name 
               AND tc.constraint_type = 'PRIMARY KEY') AS is_pk,
            -- FK
            (SELECT count(*) > 0 
             FROM information_schema.table_constraints tc
             JOIN information_schema.key_column_usage kcu ON tc.constraint_name = kcu.constraint_name
             WHERE tc.table_schema = c.table_schema 
               AND tc.table_name = c.table_name 
               AND kcu.column_name = c.column_name 
               AND tc.constraint_type = 'FOREIGN KEY') AS is_fk,
            -- FK Ref
            COALESCE(
                (SELECT ccu.table_name || '.' || ccu.column_name
                 FROM information_schema.table_constraints tc
                 JOIN information_schema.key_column_usage kcu ON tc.constraint_name = kcu.constraint_name
                 JOIN information_schema.constraint_column_usage ccu ON ccu.constraint_name = tc.constraint_name
                 WHERE tc.table_schema = c.table_schema 
                   AND tc.table_name = c.table_name 
                   AND kcu.column_name = c.column_name 
                   AND tc.constraint_type = 'FOREIGN KEY'
                 LIMIT 1), '') AS fk_ref,
            -- UNIQUE
            (SELECT count(*) > 0 
             FROM information_schema.table_constraints tc
             JOIN information_schema.key_column_usage kcu ON tc.constraint_name = kcu.constraint_name
             WHERE tc.table_schema = c.table_schema 
               AND tc.table_name = c.table_name 
               AND kcu.column_name = c.column_name 
               AND tc.constraint_type = 'UNIQUE') AS is_unique,
            -- Comment
            COALESCE(pg_catalog.col_description(pgc.oid, c.ordinal_position), '') AS description
        FROM 
            information_schema.columns c
        JOIN 
            pg_catalog.pg_class pgc ON pgc.relname = c.table_name
        JOIN 
            pg_catalog.pg_namespace pgn ON pgn.oid = pgc.relnamespace AND pgn.nspname = c.table_schema
        WHERE 
            c.table_schema = 'public'
            AND c.table_name = %s
        ORDER BY 
            c.ordinal_position;
    """

    diccionario_metadatos = {}
    for mod_name, tablas in MODULOS.items():
        for t in tablas:
            if t in tablas_existentes:
                cur.execute(query_columnas, (t,))
                columnas = []
                for row in cur.fetchall():
                    columnas.append({
                        'name': row[0],
                        'type': row[1],
                        'char_len': row[2],
                        'num_prec': row[3],
                        'num_scale': row[4],
                        'nullable': row[5],
                        'default_val': row[6],
                        'is_pk': row[7],
                        'is_fk': row[8],
                        'fk_ref': row[9],
                        'is_unique': row[10],
                        'description': row[11]
                    })
                diccionario_metadatos[t] = columnas

    cur.close()
    conn.close()
    return diccionario_metadatos

def generar_excel_por_bloques(diccionario_metadatos):
    wb = Workbook()
    
    # Configuración de estilos
    color_primary = "1B365D"       # Azul Marino UAGRM para cabecera de la tabla
    color_accent = "4A90E2"        # Azul para Módulos
    color_light_gray = "F5F5F5"    # Relleno sutil para datos
    color_pk = "E2EFDA"            # Verde pastel para PK
    color_fk = "FFF2CC"            # Amarillo pastel para FK
    
    font_modulo = Font(name="Segoe UI", size=13, bold=True, color="1B365D")
    font_tabla_name = Font(name="Segoe UI", size=11, bold=True, color="333333")
    font_tabla_desc = Font(name="Segoe UI", size=10, italic=True, color="555555")
    
    font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10)
    font_body_bold = Font(name="Segoe UI", size=10, bold=True)
    
    fill_header = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
    fill_pk = PatternFill(start_color=color_pk, end_color=color_pk, fill_type="solid")
    fill_fk = PatternFill(start_color=color_fk, end_color=color_fk, fill_type="solid")
    fill_zebra = PatternFill(start_color=color_light_gray, end_color=color_light_gray, fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    
    ws = wb.active
    ws.title = "Diccionario por Tablas"
    ws.views.sheetView[0].showGridLines = True
    
    current_row = 2
    
    # Recorrer módulos y tablas
    for modulo, tablas in MODULOS.items():
        # Escribir el título del Módulo (solo si tiene tablas creadas en la base de datos)
        tablas_validas = [t for t in tablas if t in diccionario_metadatos]
        if not tablas_validas:
            continue
            
        ws.cell(row=current_row, column=1, value=f"Módulo: {modulo}").font = font_modulo
        ws.row_dimensions[current_row].height = 24
        current_row += 2 # Espacio
        
        for idx, t in enumerate(tablas_validas, start=1):
            # Nombre de la tabla
            ws.cell(row=current_row, column=1, value=f"Tabla {idx}: {t}").font = font_tabla_name
            ws.row_dimensions[current_row].height = 20
            current_row += 1
            
            # Descripción de la tabla
            desc_tabla = TABLE_DESCRIPTIONS.get(t, f"Almacena información sobre {t}.")
            ws.cell(row=current_row, column=1, value=f"Descripción: {desc_tabla}").font = font_tabla_desc
            ws.row_dimensions[current_row].height = 18
            current_row += 2 # Espacio antes de la tabla
            
            # Cabecera de la Tabla
            headers = [
                "Nombre del Campo", 
                "Tipo de Dato (PostgreSQL)", 
                "Longitud / Precisión", 
                "Restricciones / Claves", 
                "Descripción"
            ]
            
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=h)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
                cell.border = thin_border
                
            ws.row_dimensions[current_row].height = 26
            current_row += 1
            
            # Filas de la tabla
            columnas = diccionario_metadatos[t]
            for row_idx, col in enumerate(columnas):
                # 1. Nombre del campo
                c_name = ws.cell(row=current_row, column=1, value=col['name'])
                c_name.font = font_body_bold
                c_name.border = thin_border
                c_name.alignment = align_left
                
                # 2. Tipo de Dato (PostgreSQL)
                # Formatear el tipo de dato
                raw_type = col['type'].upper()
                # Simplificación de tipos
                if "CHARACTER VARYING" in raw_type or "VARCHAR" in raw_type:
                    pg_type = "VARCHAR"
                elif "INTEGER" in raw_type or "INT" in raw_type:
                    # Si tiene un default que empieza con nextval, es SERIAL
                    if col['default_val'] and col['default_val'].startswith("nextval"):
                        pg_type = "SERIAL"
                    else:
                        pg_type = "INTEGER"
                elif "BIGINT" in raw_type:
                    if col['default_val'] and col['default_val'].startswith("nextval"):
                        pg_type = "BIGSERIAL"
                    else:
                        pg_type = "BIGINT"
                elif "TIMESTAMP" in raw_type:
                    pg_type = "TIMESTAMP"
                elif "BOOLEAN" in raw_type or "BOOL" in raw_type:
                    pg_type = "BOOLEAN"
                else:
                    pg_type = raw_type
                    
                c_type = ws.cell(row=current_row, column=2, value=pg_type)
                c_type.font = font_body
                c_type.border = thin_border
                c_type.alignment = align_center
                
                # 3. Longitud / Precisión
                if col['char_len'] is not None:
                    length_val = str(col['char_len'])
                elif col['num_prec'] is not None and col['num_scale'] is not None:
                    length_val = f"{col['num_prec']},{col['num_scale']}"
                else:
                    length_val = "-"
                    
                c_len = ws.cell(row=current_row, column=3, value=length_val)
                c_len.font = font_body
                c_len.border = thin_border
                c_len.alignment = align_center
                
                # 4. Restricciones / Claves
                restricciones = []
                
                # PK/FK
                if col['is_pk']:
                    restricciones.append("PK (Primary Key)")
                if col['is_fk']:
                    ref_tbl = col['fk_ref'].split('.')[0] if '.' in col['fk_ref'] else col['fk_ref']
                    restricciones.append(f"FK (Foreign Key) references {ref_tbl}")
                
                # Nullability
                if col['nullable'] == 'NO':
                    # Si ya es PK, está implícito, pero lo agregamos o no según preferencia.
                    restricciones.append("NOT NULL")
                else:
                    restricciones.append("NULL (Opcional)")
                    
                # Unique
                if col['is_unique']:
                    restricciones.append("UNIQUE")
                    
                # Default
                if col['default_val'] is not None:
                    # Limpiar valor default
                    def_val = col['default_val']
                    if "nextval" not in def_val:
                        # quitar casts ::character varying, etc.
                        if "::" in def_val:
                            def_val = def_val.split("::")[0]
                        restricciones.append(f"DEFAULT {def_val}")
                
                restriccion_texto = ", ".join(restricciones)
                c_rest = ws.cell(row=current_row, column=4, value=restriccion_texto)
                c_rest.font = font_body
                c_rest.border = thin_border
                c_rest.alignment = align_left
                
                # Coloración según clave
                if col['is_pk']:
                    c_rest.fill = fill_pk
                    c_rest.font = font_body_bold
                elif col['is_fk']:
                    c_rest.fill = fill_fk
                    c_rest.font = font_body_bold
                elif row_idx % 2 == 1: # Zebra striping
                    c_rest.fill = fill_zebra
                    
                # Zebra striping para el resto de celdas
                if row_idx % 2 == 1:
                    c_name.fill = fill_zebra
                    c_type.fill = fill_zebra
                    c_len.fill = fill_zebra
                    
                # 5. Descripción
                # Si está vacía en PostgreSQL, autogenerarla de forma inteligente
                desc_text = col['description']
                if not desc_text or desc_text.strip() == "" or desc_text == "Pendiente de descripción técnica":
                    desc_text = generar_descripcion_campo(t, col['name'])
                    
                c_desc = ws.cell(row=current_row, column=5, value=desc_text)
                c_desc.font = font_body
                c_desc.border = thin_border
                c_desc.alignment = align_left
                
                if row_idx % 2 == 1:
                    c_desc.fill = fill_zebra
                
                ws.row_dimensions[current_row].height = 20
                current_row += 1
                
            current_row += 3 # Espacio para la siguiente tabla
            
        current_row += 1 # Espacio adicional entre módulos

    # Ajustar ancho de las columnas a un tamaño cómodo
    ws.column_dimensions['A'].width = 25  # Nombre del Campo
    ws.column_dimensions['B'].width = 28  # Tipo de Dato
    ws.column_dimensions['C'].width = 22  # Longitud / Precisión
    ws.column_dimensions['D'].width = 38  # Restricciones / Claves
    ws.column_dimensions['E'].width = 65  # Descripción (ancho extra para textos largos)

    filename = "diccionario_datos_por_tablas.xlsx"
    wb.save(filename)
    return filename

if __name__ == "__main__":
    print("Iniciando generación de diccionario de datos agrupado por tablas...")
    metadata = obtener_metadatos_base_datos()
    if metadata:
        archivo = generar_excel_por_bloques(metadata)
        print(f"¡Éxito! Diccionario de datos en bloque generado en: {os.path.abspath(archivo)}")
    else:
        print("Error al obtener los metadatos de la base de datos.")
