import os
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Configuración de conexión a PostgreSQL
DB_CONFIG = {
    'dbname': 'bdexpo',
    'user': 'postgres',
    'password': 'root',
    'host': 'localhost',
    'port': '5432'
}

def obtener_metadatos():
    try:
        conn = psycopg2.connect(**DB_CONFIG)
        cur = conn.cursor()
    except Exception as e:
        print(f"Error al conectar a la base de datos: {e}")
        print("Por favor, asegúrate de que PostgreSQL está activo y los datos de conexión en el script son correctos.")
        return None, None

    # 1. Obtener todas las tablas del esquema público
    query_tablas = """
        SELECT table_name
        FROM information_schema.tables
        WHERE table_schema = 'public'
          AND table_type = 'BASE TABLE'
        ORDER BY table_name;
    """
    cur.execute(query_tablas)
    tablas = [r[0] for r in cur.fetchall()]

    # Filtros de tablas de Django y sistema
    tablas_negocio = []
    tablas_sistema = []
    
    apps_prefixes = ('usuarios', 'academico', 'eventos', 'proyectos', 'evaluaciones', 'premiacion')
    
    for t in tablas:
        # Si empieza con alguno de los prefijos de las apps, es de negocio
        if t.startswith(apps_prefixes):
            tablas_negocio.append(t)
        # Tablas especiales que no tienen el prefijo directo pero son de negocio
        elif t in ('usuario', 'participante', 'tutor', 'tribunal', 'personal', 'proyecto', 'premio', 'cronograma', 'cargo', 'carrera', 'ea_carrera', 'membrete_firmante', 'tribunal_ext'):
            tablas_negocio.append(t)
        else:
            tablas_sistema.append(t)

    # Ordenar ambas listas
    tablas_negocio.sort()
    tablas_sistema.sort()

    diccionario_completo = {}
    
    # 2. Obtener columnas por cada tabla
    query_columnas = """
        SELECT 
            c.column_name,
            c.data_type,
            COALESCE(c.character_maximum_length::text, 
                     CASE 
                         WHEN c.numeric_precision IS NOT NULL AND c.numeric_scale IS NOT NULL 
                         THEN c.numeric_precision || ',' || c.numeric_scale
                         ELSE ''
                     END) AS length_precision,
            c.is_nullable,
            (SELECT count(*) > 0 
             FROM information_schema.table_constraints tc
             JOIN information_schema.key_column_usage kcu ON tc.constraint_name = kcu.constraint_name
             WHERE tc.table_schema = c.table_schema 
               AND tc.table_name = c.table_name 
               AND kcu.column_name = c.column_name 
               AND tc.constraint_type = 'PRIMARY KEY') AS is_pk,
            (SELECT count(*) > 0 
             FROM information_schema.table_constraints tc
             JOIN information_schema.key_column_usage kcu ON tc.constraint_name = kcu.constraint_name
             WHERE tc.table_schema = c.table_schema 
               AND tc.table_name = c.table_name 
               AND kcu.column_name = c.column_name 
               AND tc.constraint_type = 'FOREIGN KEY') AS is_fk,
            COALESCE(
                (SELECT ccu.table_name || '.' || ccu.column_name
                 FROM information_schema.table_constraints tc
                 JOIN information_schema.key_column_usage kcu ON tc.constraint_name = kcu.constraint_name
                 JOIN information_schema.constraint_column_usage ccu ON ccu.constraint_name = tc.constraint_name
                 WHERE tc.table_schema = c.table_schema 
                   AND tc.table_name = c.table_name 
                   AND kcu.column_name = c.column_name 
                   AND tc.constraint_type = 'FOREIGN KEY'
                 LIMIT 1), '') AS fk_ref,
            COALESCE(pg_catalog.col_description(pgc.oid, c.ordinal_position), '') AS description
        FROM 
            information_schema.columns c
        JOIN 
            pg_catalog.pg_class pgc ON pgc.relname = c.table_name
        JOIN 
            pg_catalog.pg_namespace pgn ON pgn.oid = pgc.relnamespace AND pgn.nspname = c.table_schema
        WHERE 
            c.table_schema = 'public'
            AND c.table_name = %s
        ORDER BY 
            c.ordinal_position;
    """

    for t in tablas:
        cur.execute(query_columnas, (t,))
        columnas = []
        for row in cur.fetchall():
            columnas.append({
                'name': row[0],
                'type': row[1],
                'length': row[2],
                'nullable': 'Sí' if row[3] == 'YES' else 'No',
                'is_pk': row[4],
                'is_fk': row[5],
                'fk_ref': row[6],
                'description': row[7]
            })
        diccionario_completo[t] = columnas

    cur.close()
    conn.close()

    return {
        'negocio': tablas_negocio,
        'sistema': tablas_sistema
    }, diccionario_completo

def generar_excel(tablas_clasificadas, diccionario):
    wb = Workbook()
    
    # ----------------------------------------------------
    # ESTILOS GENERALES Y PALETA DE COLORES
    # ----------------------------------------------------
    # Paleta Azul UAGRM Profesional
    color_primary = "1B365D"       # Azul Marino Oscuro para cabeceras principales
    color_accent = "4A90E2"        # Azul Brillante para sub-cabeceras / decoraciones
    color_zebra = "F5F8FC"         # Azul Grisáceo muy claro para filas alternas
    color_pk_fill = "E2EFDA"       # Verde claro para Llaves Primarias
    color_fk_fill = "FFF2CC"       # Amarillo claro para Llaves Foráneas
    color_title_bg = "0D233A"      # Azul ultra oscuro para la portada/título
    
    # Fuentes
    font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="D9E1F2")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10)
    font_body_bold = Font(name="Segoe UI", size=10, bold=True)
    font_table_title = Font(name="Segoe UI", size=12, bold=True, color="1B365D")
    
    # Fills (Rellenos)
    fill_title = PatternFill(start_color=color_title_bg, end_color=color_title_bg, fill_type="solid")
    fill_header = PatternFill(start_color=color_primary, end_color=color_primary, fill_type="solid")
    fill_zebra = PatternFill(start_color=color_zebra, end_color=color_zebra, fill_type="solid")
    fill_pk = PatternFill(start_color=color_pk_fill, end_color=color_pk_fill, fill_type="solid")
    fill_fk = PatternFill(start_color=color_fk_fill, end_color=color_fk_fill, fill_type="solid")
    
    # Bordes
    thin_border_side = Side(border_style="thin", color="D3D3D3")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    double_bottom = Border(bottom=Side(border_style="double", color="1B365D"))
    
    # Alineación
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_title = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ----------------------------------------------------
    # HOJA 1: PORTADA E ÍNDICE DE TABLAS
    # ----------------------------------------------------
    ws_index = wb.active
    ws_index.title = "Índice de Tablas"
    ws_index.views.sheetView[0].showGridLines = True
    
    # Encabezado de la Portada
    ws_index.merge_cells("A1:E2")
    ws_index["A1"] = "DICCIONARIO DE DATOS - SISTEMA EXPOCIENCIA UAGRM"
    ws_index["A1"].font = font_title
    ws_index["A1"].alignment = align_title
    ws_index["A1"].fill = fill_title
    
    ws_index["A3"] = "Este documento contiene la estructura detallada de la base de datos local 'bdexpo' del sistema."
    ws_index["A3"].font = font_subtitle
    ws_index["A3"].fill = PatternFill(start_color="1A3052", end_color="1A3052", fill_type="solid")
    ws_index.merge_cells("A3:E3")
    
    ws_index.row_dimensions[1].height = 25
    ws_index.row_dimensions[2].height = 25
    ws_index.row_dimensions[3].height = 20
    ws_index.row_dimensions[4].height = 15 # Fila en blanco
    
    # Tabla de Índice
    ws_index["A5"] = "ÍNDICE DE TABLAS DEL SISTEMA"
    ws_index["A5"].font = font_table_title
    
    headers_index = ["Nº", "Nombre de la Tabla", "Módulo / Aplicación", "Tipo de Tabla", "Nº Columnas"]
    for col_idx, h in enumerate(headers_index, start=1):
        cell = ws_index.cell(row=6, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
    
    ws_index.row_dimensions[6].height = 28
    
    row_num = 7
    count = 1
    
    # Escribir tablas de negocio
    for t in tablas_clasificadas['negocio']:
        ws_index.cell(row=row_num, column=1, value=count).alignment = align_center
        
        # Nombre de la tabla
        cell_t = ws_index.cell(row=row_num, column=2, value=t)
        cell_t.font = font_body_bold
        
        # Módulo sugerido
        parts = t.split('_')
        modulo = parts[0].upper() if len(parts) > 1 else "GENERAL"
        ws_index.cell(row=row_num, column=3, value=modulo).alignment = align_center
        
        ws_index.cell(row=row_num, column=4, value="Negocio (Expociencia)").alignment = align_center
        
        num_cols = len(diccionario.get(t, []))
        ws_index.cell(row=row_num, column=5, value=num_cols).alignment = align_center
        
        # Formatos comunes de fila
        for col_idx in range(1, 6):
            cell = ws_index.cell(row=row_num, column=col_idx)
            if col_idx != 2:
                cell.font = font_body
            cell.border = border_all
            if row_num % 2 == 0:
                cell.fill = fill_zebra
                
        ws_index.row_dimensions[row_num].height = 20
        row_num += 1
        count += 1
        
    # Escribir tablas de sistema
    for t in tablas_clasificadas['sistema']:
        ws_index.cell(row=row_num, column=1, value=count).alignment = align_center
        
        cell_t = ws_index.cell(row=row_num, column=2, value=t)
        cell_t.font = font_body_bold
        
        ws_index.cell(row=row_num, column=3, value="SISTEMA / DJANGO").alignment = align_center
        ws_index.cell(row=row_num, column=4, value="Interno (Framework)").alignment = align_center
        
        num_cols = len(diccionario.get(t, []))
        ws_index.cell(row=row_num, column=5, value=num_cols).alignment = align_center
        
        for col_idx in range(1, 6):
            cell = ws_index.cell(row=row_num, column=col_idx)
            if col_idx != 2:
                cell.font = font_body
            cell.border = border_all
            if row_num % 2 == 0:
                cell.fill = fill_zebra
                
        ws_index.row_dimensions[row_num].height = 20
        row_num += 1
        count += 1

    # Ajustar anchos del índice
    for col in ws_index.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Evitar celdas combinadas del título en la fila 1-3
            if cell.row <= 4:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws_index.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # ----------------------------------------------------
    # HOJA 2: DICCIONARIO DE DATOS DETALLADO
    # ----------------------------------------------------
    ws_dict = wb.create_sheet(title="Diccionario de Datos")
    ws_dict.views.sheetView[0].showGridLines = True
    
    headers_dict = [
        "Tabla", "Campo / Columna", "Tipo de Dato", 
        "Longitud", "Nulos", "PK", "FK", "Referencia (FK)", "Descripción / Propósito"
    ]
    
    # Escribir cabeceras
    for col_idx, h in enumerate(headers_dict, start=1):
        cell = ws_dict.cell(row=1, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all
        
    ws_dict.row_dimensions[1].height = 28
    
    dict_row = 2
    
    # Escribir todas las columnas ordenadas por tabla (primero negocio, luego sistema)
    todas_las_tablas = tablas_clasificadas['negocio'] + tablas_clasificadas['sistema']
    
    last_table = None
    
    for t in todas_las_tablas:
        cols_data = diccionario.get(t, [])
        for col in cols_data:
            # Si cambia de tabla, podemos poner una línea divisoria o dejar el nombre de la tabla
            ws_dict.cell(row=dict_row, column=1, value=t).font = font_body_bold
            
            cell_name = ws_dict.cell(row=dict_row, column=2, value=col['name'])
            cell_name.font = font_body_bold
            
            ws_dict.cell(row=dict_row, column=3, value=col['type'])
            ws_dict.cell(row=dict_row, column=4, value=col['length']).alignment = align_center
            ws_dict.cell(row=dict_row, column=5, value=col['nullable']).alignment = align_center
            
            # PK
            pk_val = "PK" if col['is_pk'] else ""
            cell_pk = ws_dict.cell(row=dict_row, column=6, value=pk_val)
            cell_pk.alignment = align_center
            if col['is_pk']:
                cell_pk.fill = fill_pk
                cell_pk.font = font_body_bold
                
            # FK
            fk_val = "FK" if col['is_fk'] else ""
            cell_fk = ws_dict.cell(row=dict_row, column=7, value=fk_val)
            cell_fk.alignment = align_center
            if col['is_fk']:
                cell_fk.fill = fill_fk
                cell_fk.font = font_body_bold
                
            # FK Ref
            ws_dict.cell(row=dict_row, column=8, value=col['fk_ref'])
            
            # Descripción
            # Si está vacía en Postgres, podemos poner un placeholder sutil para que el usuario la llene
            desc_val = col['description'] if col['description'] else "Pendiente de descripción técnica"
            cell_desc = ws_dict.cell(row=dict_row, column=9, value=desc_val)
            if not col['description']:
                cell_desc.font = Font(name="Segoe UI", size=10, italic=True, color="8C8C8C")
            
            # Estilos de celda
            for col_idx in range(1, 10):
                cell = ws_dict.cell(row=dict_row, column=col_idx)
                if col_idx in (1, 2) or (col_idx == 6 and col['is_pk']) or (col_idx == 7 and col['is_fk']):
                    cell.font = font_body_bold
                else:
                    if col_idx == 9 and not col['description']:
                        pass
                    else:
                        cell.font = font_body
                cell.border = border_all
                # Zebra striping agrupado por tabla para mejor legibilidad
                if last_table and last_table != t:
                    pass # Podríamos poner una línea gruesa
                
                # Zebra striping normal alterno por fila
                if dict_row % 2 == 0 and not cell.fill.fill_type:
                    cell.fill = fill_zebra
            
            ws_dict.row_dimensions[dict_row].height = 20
            dict_row += 1
            
        last_table = t

    # Habilitar autofiltros en el diccionario de datos
    ws_dict.auto_filter.ref = f"A1:I{dict_row - 1}"

    # Ajustar anchos del diccionario
    for col in ws_dict.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                # Evitar que descripciones muy largas hagan columnas excesivamente anchas
                val_str = str(cell.value)
                if cell.column == 9 and len(val_str) > 40:
                    max_len = max(max_len, 40)
                else:
                    max_len = max(max_len, len(val_str))
        ws_dict.column_dimensions[col_letter].width = max(max_len + 4, 10)

    # Guardar archivo
    output_filename = "diccionario_datos_expociencia.xlsx"
    wb.save(output_filename)
    return output_filename

if __name__ == "__main__":
    print("Iniciando la extracción de metadatos de la base de datos 'bdexpo'...")
    tablas, diccionario = obtener_metadatos()
    if tablas and diccionario:
        print(f"Metadatos obtenidos: {len(tablas['negocio'])} tablas de negocio y {len(tablas['sistema'])} tablas del sistema.")
        print("Generando archivo Excel formateado profesionalmente...")
        filename = generar_excel(tablas, diccionario)
        print(f"¡Éxito! El diccionario de datos ha sido generado y guardado en: {os.path.abspath(filename)}")
    else:
        print("No se pudo generar el diccionario de datos debido a un error.")
